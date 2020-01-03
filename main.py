# import openpyxl and tkinter modules
from openpyxl import *
from tkinter import *

# globally declare wb and sheet variable

# opening the existing excel file
wb = load_workbook('excel.xlsx')

# create the sheet object
sheet = wb.active


def excel():
    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 10

    # write given data to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "Name of Product"
    sheet.cell(row=1, column=2).value = "Quantity in grams"
    sheet.cell(row=1, column=3).value = "No. of pieces"
    sheet.cell(row=1, column=4).value = "MRP"


# Function to set focus
def focus1(event):
    # set focus on the mrp box
    name_field.focus_set()


# Function to set focus (cursor)
def focus2(event):
    # set focus on the quantity_grams box
    quantity_grams.focus_set()


# Function to set focus
def focus3(event):
    # set focus on the no_of_pieces box
    no_of_pieces.focus_set()


# Function to set focus
def focus4(event):
    # set focus on the mrp box
    mrp.focus_set()


# Function for clearing the
# contents of text entry boxes
def clear():
    # clear the content of text entry box
    name_field.delete(0, END)
    quantity_grams.delete(0, END)
    no_of_pieces.delete(0, END)
    mrp.delete(0, END)


# Function to take data from GUI
# window and write to an excel file
def insert():
    # if user not fill any entry
    # then print "empty input"
    if (name_field.get() == "" and
            quantity_grams.get() == "" and
            no_of_pieces.get() == "" and
            mrp.get() == ""):

        print("empty input")

    else:

        # assigning the max row and max column
        # value upto which data is written
        # in an excel sheet to the variable
        current_row = sheet.max_row
        current_column = sheet.max_column

        # get method returns current text
        # as string which we write into
        # excel spreadsheet at particular location
        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = quantity_grams.get()
        sheet.cell(row=current_row + 1, column=3).value = no_of_pieces.get()
        sheet.cell(row=current_row + 1, column=4).value = mrp.get()

        # save the file
        wb.save('excel.xlsx')

        # set focus on the name_field box
        name_field.focus_set()

        # call the clear() function
        clear()

    # Driver code


if __name__ == "__main__":
    # create a GUI window
    root = Tk()

    # set the background colour of GUI window
    root.configure(background='light green')

    # set the title of GUI window
    root.title("Quantity form")

    # set the configuration of GUI window
    root.geometry("500x300")

    excel()

    # create a Form label
    heading = Label(root, text="Form", bg="light green")

    # create a Name label
    name = Label(root, text="Name of Product", bg="light green")

    # create a Course label
    quantity = Label(root, text="Quantity in grams", bg="light green")

    # create a Semester label
    nums = Label(root, text="No. of pieces", bg="light green")

    # create a Form No. lable
    mrp_var = Label(root, text="MRP", bg="light green")

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    heading.grid(row=0, column=1)
    name.grid(row=1, column=0)
    quantity.grid(row=2, column=0)
    nums.grid(row=3, column=0)
    mrp_var.grid(row=4, column=0)

    # create a text entry box
    # for typing the information
    name_field = Entry(root)
    quantity_grams = Entry(root)
    no_of_pieces = Entry(root)
    mrp = Entry(root)

    # bind method of widget is used for
    # the binding the function with the events

    # whenever the enter key is pressed
    # then call the focus1 function
    name_field.bind("<Return>", focus1)

    # whenever the enter key is pressed
    # then call the focus2 function
    quantity_grams.bind("<Return>", focus2)

    # whenever the enter key is pressed
    # then call the focus3 function
    no_of_pieces.bind("<Return>", focus3)

    # whenever the enter key is pressed
    # then call the focus4 function
    mrp.bind("<Return>", focus4)

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    name_field.grid(row=1, column=1, ipadx="100")
    quantity_grams.grid(row=2, column=1, ipadx="100")
    no_of_pieces.grid(row=3, column=1, ipadx="100")
    mrp.grid(row=4, column=1, ipadx="100")

    # call excel function
    excel()

    # create a Submit Button and place into the root window
    submit = Button(root, text="Submit", fg="Black",
                    bg="Red", command=insert)
    submit.grid(row=8, column=1)

    # start the GUI
    root.mainloop()
